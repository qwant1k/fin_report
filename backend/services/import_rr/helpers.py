"""Утилиты для импортёра Risk Report XLSM."""
from __future__ import annotations

import hashlib
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl.worksheet.worksheet import Worksheet


class _MatCell:
    """Лёгкая ячейка: имеет только атрибут ``.value`` (как у openpyxl Cell)."""
    __slots__ = ("value",)

    def __init__(self, value: Any):
        self.value = value


class MaterializedSheet:
    """Снимок листа в виде 2D-списка значений.

    Полностью совместим с подмножеством API ``openpyxl.Worksheet``, которое
    используют наши парсеры: ``cell(row, col).value``, ``max_row``,
    ``max_column``, ``title``, итерация по строкам.

    Зачем: в режиме ``read_only=True`` openpyxl возвращает «съехавшие» данные
    при многократных вызовах ``ws.cell(r, c)`` поверх нескольких листов одной
    книги (внутренний потоковый парсер делит общее состояние). Один проход
    ``iter_rows(values_only=True)`` возвращает корректные значения, после
    чего мы работаем уже с in-memory снимком.
    """
    __slots__ = ("title", "_data", "max_row", "max_column")

    def __init__(self, title: str, rows: list[tuple]):
        self.title = title
        self._data = rows  # list[tuple] — индекс 0 = excel row 1
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)

    def cell(self, row: int, col: int) -> _MatCell:
        """1-индексный доступ к ячейке. Возвращает заглушку с ``.value``."""
        if row < 1 or col < 1 or row > self.max_row:
            return _MatCell(None)
        rec = self._data[row - 1]
        if col > len(rec):
            return _MatCell(None)
        return _MatCell(rec[col - 1])

    def iter_rows(self, *, min_row: int = 1, max_row: Optional[int] = None,
                  values_only: bool = True, **_kwargs):
        """Совместимая итерация (только values_only=True)."""
        end = max_row or self.max_row
        for r in range(min_row, end + 1):
            rec = self._data[r - 1] if r - 1 < len(self._data) else ()
            if values_only:
                yield tuple(rec)
            else:
                yield tuple(_MatCell(v) for v in rec)


def materialize_sheet(ws: Worksheet) -> MaterializedSheet:
    """Снять снимок листа openpyxl в ``MaterializedSheet`` за один проход.

    Безопасно использовать в read_only-режиме: ``iter_rows(values_only=True)``
    идемпотентен и не страдает от стейт-блида между листами.
    """
    rows: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row if isinstance(row, tuple) else tuple(row))
    return MaterializedSheet(getattr(ws, "title", ""), rows)

# Шаблон имени файла: risk report_DDMMYYYY_.xlsm  (с подчёркиванием в конце или без)
RR_FILENAME_RE = re.compile(
    r"risk[\s_]*report[\s_]*[-_]?(\d{2})\.?(\d{2})\.?(\d{4})_?",
    re.IGNORECASE,
)


def extract_date_from_filename(file_path: Path | str) -> Optional[date]:
    """Достать дату из имени файла Risk Report.
    Поддерживает: risk report_20102025_.xlsm, risk report 20.10.2025.xlsm и т.п."""
    name = Path(file_path).stem
    m = RR_FILENAME_RE.search(name)
    if not m:
        # Альтернативный шаблон: risk_report_20251020.xlsm
        m2 = re.search(r"(\d{8})", name)
        if m2:
            try:
                s = m2.group(1)
                return date(int(s[4:8]), int(s[2:4]), int(s[0:2]))
            except (ValueError, IndexError):
                pass
        return None
    try:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return date(y, mo, d)
    except ValueError:
        return None


def file_sha256(path: Path) -> str:
    """SHA-256 содержимого файла (для дедупликации источника)."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 16), b""):
            h.update(chunk)
    return h.hexdigest()


def cell_str(value: Any) -> Optional[str]:
    """Безопасное приведение ячейки к str с обрезкой пробелов."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s and s != "-" else None


def cell_float(value: Any) -> Optional[float]:
    """Распознать число из ячейки (поддерживает каз. формат пробел+запятая)."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return f if f == f else None  # NaN check
    s = str(value).strip().replace("\u00A0", " ")
    if not s or s == "-":
        return None
    has_comma = "," in s
    has_dot = "." in s
    if has_comma and has_dot:
        # KASE XLSX: "1,214.9000" means 1214.9; comma is thousands separator.
        s = s.replace(",", "").replace(" ", "")
    elif has_comma:
        # Local format: "1 214,90" means 1214.9; comma is decimal separator.
        s = s.replace(" ", "").replace(",", ".")
    else:
        s = s.replace(" ", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def cell_int(value: Any) -> Optional[int]:
    f = cell_float(value)
    return int(round(f)) if f is not None else None


def cell_date(value: Any) -> Optional[date]:
    """Распознать дату — datetime/date/строка ДД.ММ.ГГГГ или ГГГГ-ММ-ДД."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s or s == "-":
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y.%m.%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def normalize_header(s: Any) -> str:
    """Шапка → нижний регистр без множ. пробелов и табов."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def find_header_row(ws: Worksheet, *, max_scan: int = 30,
                    must_contain: Iterable[str] = ()) -> Optional[int]:
    """Найти строку с шапкой колонок: первая строка, где встречаются все ключевые слова.

    `must_contain` — нижний регистр; ищется как подстрока в склейке всех ячеек строки.
    """
    needed = [n.lower() for n in must_contain]
    if not needed:
        return None
    upper = min(ws.max_row or 0, max_scan)
    for r in range(1, upper + 1):
        joined = " | ".join(
            normalize_header(ws.cell(r, c).value) for c in range(1, (ws.max_column or 0) + 1)
        )
        if all(k in joined for k in needed):
            return r
    return None


def header_index_map(ws: Worksheet, header_row: int,
                     *, max_col: Optional[int] = None) -> dict[str, int]:
    """Маппинг нормализованной шапки → номер колонки (1-индексный)."""
    out: dict[str, int] = {}
    upper = max_col if max_col is not None else (ws.max_column or 0)
    for c in range(1, upper + 1):
        h = normalize_header(ws.cell(header_row, c).value)
        if h and h not in out:
            out[h] = c
    return out


def find_col(headers: dict[str, int], *aliases: str) -> Optional[int]:
    """Найти колонку по любому из псевдонимов (case-insensitive substring match)."""
    for alias in aliases:
        a = alias.lower()
        if a in headers:
            return headers[a]
    # substring fallback
    for h, c in headers.items():
        for alias in aliases:
            if alias.lower() in h:
                return c
    return None


def iter_data_rows(
    ws: Worksheet,
    start_row: int,
    *,
    stop_on_empty_a: bool = True,
    key_col: int = 1,
    max_empty: int = 5,
):
    """Итератор по строкам данных начиная со ``start_row``.

    По умолчанию проверяет колонку A (исторический контракт), но в Risk Report
    некоторые листы имеют пустую колонку A (например, ГЦБ/Агентские/МФО — данные
    начинаются с колонки B, ISIN). В таких случаях передайте ``key_col``,
    указывающий на колонку, по которой определяется наличие данных.
    """
    empty_streak = 0
    for r in range(start_row, (ws.max_row or 0) + 1):
        first = ws.cell(r, key_col).value
        if first is None or (isinstance(first, str) and not first.strip()):
            empty_streak += 1
            if stop_on_empty_a and empty_streak >= max_empty:
                break
            continue
        empty_streak = 0
        yield r


def normalize_subfund_name(s: Any) -> Optional[str]:
    """Канонизация Sub portfolio name → имя CDU из словаря CDU_NAME_ALIASES."""
    from services.calculator.constants import normalize_cdu_name
    return normalize_cdu_name(cell_str(s))


def safe_sheet(wb, *names: str):
    """Получить лист по любому из имён или None.

    Алгоритм (приоритет от точного к нечёткому, чтобы избежать ложных срабатываний
    типа ``Repo`` ⊂ ``Report``):

    1. Точное совпадение нормализованных имён (без пробелов/`_-:.()`).
    2. Целевое имя — префикс имени листа (``repo`` → ``repo_lots``).
    3. Имя листа — префикс целевого (``mbm index`` → ``mbm index - с 1 апр…``).
    4. Двусторонняя подстрока — последний фоллбэк.
    """
    if not wb:
        return None
    targets = [re.sub(r"[\s_:.()-]+", "", n.lower()) for n in names if n]
    if not targets:
        return None
    sheets = [(s, re.sub(r"[\s_:.()-]+", "", s.lower())) for s in wb.sheetnames]

    # Pass 1 — exact
    for t in targets:
        for sheet_name, normed in sheets:
            if t == normed:
                return wb[sheet_name]
    # Pass 2 — target — префикс имени листа
    for t in targets:
        for sheet_name, normed in sheets:
            if normed.startswith(t):
                return wb[sheet_name]
    # Pass 3 — имя листа — префикс target
    for t in targets:
        for sheet_name, normed in sheets:
            if t.startswith(normed) and normed:
                return wb[sheet_name]
    # Pass 4 — двусторонний substring (старое поведение)
    for t in targets:
        for sheet_name, normed in sheets:
            if t in normed or normed in t:
                return wb[sheet_name]
    return None
