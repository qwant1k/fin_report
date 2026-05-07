"""Парсеры конкретных листов Risk Report XLSM.

Каждая функция принимает worksheet и возвращает структурированный список словарей
для дальнейшего сохранения в БД. Импортёр-оркестратор вызывает их по очереди.
"""
from __future__ import annotations

from datetime import date
from typing import Any, Optional

from loguru import logger
from openpyxl.worksheet.worksheet import Worksheet

from .helpers import (
    cell_date,
    cell_float,
    cell_int,
    cell_str,
    find_col,
    find_header_row,
    header_index_map,
    iter_data_rows,
    normalize_subfund_name,
)



# ──────────────────────────────────────────────────────────────────────────
# Cash sheet — снимок денежных остатков (по ДУ × дата × валюта)
# ──────────────────────────────────────────────────────────────────────────
def parse_cash_sheet(ws: Worksheet, *, fallback_date: Optional[date] = None) -> list[dict]:
    """Лист `Cash`.

    Поддерживает две раскладки:

    A) **Long-format**: колонки `Date | Sub portfolio name | Currency | Amount`.
    B) **Wide-format** (как в реальных Risk Report): сабфонды разложены по
       блокам колонок, в каждом — три метрики
       ``Market Value T-1 | Cash flow | Market Value T-0``.
       Берём «сегодняшнее» значение `Market Value T-0` как остаток на дату.
    """
    # 1) Long-format
    header_row = find_header_row(ws, must_contain=("sub", "amount"))
    if header_row is None:
        header_row = find_header_row(ws, must_contain=("sub", "остаток"))
    if header_row is not None:
        headers = header_index_map(ws, header_row)
        col_date = find_col(headers, "дата", "date", "valuation date")
        col_sub = find_col(headers, "sub portfolio name", "sub-fund", "sub fund", "ду")
        col_ccy = find_col(headers, "currency", "валюта", "ccy")
        col_amt = find_col(headers, "amount", "остаток", "cash", "сумма", "value")
        if col_sub and col_amt:
            out: list[dict] = []
            for r in iter_data_rows(ws, header_row + 1):
                sub = normalize_subfund_name(ws.cell(r, col_sub).value)
                if not sub:
                    continue
                amt = cell_float(ws.cell(r, col_amt).value)
                if amt is None:
                    continue
                d = cell_date(ws.cell(r, col_date).value) if col_date else fallback_date
                if d is None:
                    d = fallback_date
                if d is None:
                    continue
                ccy = cell_str(ws.cell(r, col_ccy).value) if col_ccy else "KZT"
                out.append({
                    "snapshot_date": d,
                    "cdu_name": sub,
                    "currency": (ccy or "KZT").upper()[:8],
                    "amount": amt,
                    "src_row": r,
                })
            if out:
                return out

    # 2) Wide-format
    rows = _parse_wide_metric_sheet(
        ws,
        metric_aliases=("market value t-0", "market value t0", "t-0", "t0"),
        record_key="amount",
        defaults={"currency": "KZT"},
    )
    if rows:
        for r in rows:
            r.setdefault("snapshot_date", r.pop("date", None) or fallback_date)
        return [r for r in rows if r.get("snapshot_date") is not None]

    logger.debug("Cash sheet: header not found")
    return []


# ──────────────────────────────────────────────────────────────────────────
# MV sheet — портфельные метрики
# ──────────────────────────────────────────────────────────────────────────
def parse_mv_sheet(ws: Worksheet, *, fallback_date: Optional[date] = None) -> list[dict]:
    """Лист `MV`. Поддерживает long- и wide-format (см. ``parse_cash_sheet``).

    Wide-format: на каждый сабфонд блок ``Cash flow | Market Value | Daily return``.
    """
    # 1) Long-format
    header_row = find_header_row(ws, must_contain=("sub", "market value"))
    if header_row is not None:
        headers = header_index_map(ws, header_row)
        col_date = find_col(headers, "дата", "date", "valuation date")
        col_sub = find_col(headers, "sub portfolio name", "sub-fund", "sub fund", "ду")
        col_cf = find_col(headers, "cash flow", "денежный поток")
        col_mv = find_col(headers, "market value", "рыночная стоимость", "mv total")
        col_ret = find_col(headers, "return", "доходность")
        col_ytm = find_col(headers, "ytm", "wa-ytm", "weighted average ytm")
        col_dur = find_col(headers, "duration", "wa-duration", "weighted average duration")
        if col_sub and col_mv:
            out: list[dict] = []
            for r in iter_data_rows(ws, header_row + 1):
                sub = normalize_subfund_name(ws.cell(r, col_sub).value)
                if not sub:
                    continue
                d = cell_date(ws.cell(r, col_date).value) if col_date else fallback_date
                if d is None:
                    d = fallback_date
                if d is None:
                    continue
                out.append({
                    "snapshot_date": d,
                    "cdu_name": sub,
                    "cash_flow": cell_float(ws.cell(r, col_cf).value) if col_cf else None,
                    "market_value_total": cell_float(ws.cell(r, col_mv).value) or 0.0,
                    "return_pct": cell_float(ws.cell(r, col_ret).value) if col_ret else None,
                    "ytm_weighted": cell_float(ws.cell(r, col_ytm).value) if col_ytm else None,
                    "duration_weighted": cell_float(ws.cell(r, col_dur).value) if col_dur else None,
                    "src_row": r,
                })
            if out:
                return out

    # 2) Wide-format — несколько метрик на сабфонд
    rows = _parse_wide_metric_sheet(
        ws,
        metrics={
            "market_value_total": ("market value", "mv"),
            "cash_flow": ("cash flow",),
            "return_pct": ("daily return", "return"),
        },
        skip_metric_substrings=("t-1", "t-0", "t0"),
    )
    if rows:
        for r in rows:
            r.setdefault("snapshot_date", r.pop("date", None) or fallback_date)
            # market_value_total обязателен; нормализуем None → 0.0
            r["market_value_total"] = r.get("market_value_total") or 0.0
        return [r for r in rows if r.get("snapshot_date") is not None]

    logger.debug("MV sheet: header not found")
    return []


# ──────────────────────────────────────────────────────────────────────────
# Wide-format универсальный детектор: сабфонды как блоки колонок
# ──────────────────────────────────────────────────────────────────────────
def _parse_wide_metric_sheet(
    ws: Worksheet,
    *,
    metrics: Optional[dict[str, tuple[str, ...]]] = None,
    metric_aliases: tuple[str, ...] = (),
    record_key: Optional[str] = None,
    skip_metric_substrings: tuple[str, ...] = (),
    defaults: Optional[dict] = None,
    max_scan_rows: int = 30,
) -> list[dict]:
    """Парсер «широкой» раскладки: ``Date | <subfund1: m1, m2, m3> | <subfund2: …>``.

    Два режима:
    * ``metrics`` — словарь *имя поля → tuple алиасов*, каждая метрика выгружается
      в отдельное поле результирующего словаря (подходит для MV).
    * ``metric_aliases`` + ``record_key`` — берём ровно одну метрику и кладём её
      под именем ``record_key`` (подходит для Cash, где нужен только остаток).

    Возвращает список словарей ``{date, cdu_name, <metric fields>, src_row}``.
    """
    if metrics is None and not metric_aliases:
        return []

    # 1. Найти строку, где есть «Date» (и рядом — ряд с именами сабфондов).
    date_row, date_col = _find_date_anchor(ws, max_scan=max_scan_rows)
    if date_row is None or date_col is None:
        return []

    last_col = ws.max_column or 0

    # Имена сабфондов могут быть в той же строке, что «Date» (Cash-раскладка),
    # либо на 1–3 строки выше (MV-раскладка с двойной шапкой).
    def _scan_funds(row: int) -> list[tuple[int, str]]:
        anchors: list[tuple[int, str]] = []
        for c in range(date_col + 1, last_col + 1):
            v = cell_str(ws.cell(row, c).value)
            if not v:
                continue
            normalized = normalize_subfund_name(v)
            if normalized:
                anchors.append((c, normalized))
        return anchors

    fund_row = date_row
    fund_anchors = _scan_funds(fund_row)
    if len(fund_anchors) < 2:
        for delta in (1, 2, 3):
            candidate = date_row - delta
            if candidate < 1:
                break
            anchors = _scan_funds(candidate)
            if len(anchors) >= 2:
                fund_row = candidate
                fund_anchors = anchors
                break
    if not fund_anchors:
        return []

    # Если шапка фондов выше «Date» — метрики лежат в строке Date;
    # если шапка фондов в той же строке, что «Date» — метрики строкой ниже.
    metric_row = date_row if fund_row != date_row else date_row + 1
    data_start = max(metric_row, fund_row) + 1

    # 2. Для каждого сабфонда определить span колонок и привязать метрики.
    fund_blocks: list[tuple[int, str, dict[str, int]]] = []
    for i, (anchor_col, fund_name) in enumerate(fund_anchors):
        next_col = fund_anchors[i + 1][0] if i + 1 < len(fund_anchors) else last_col + 1
        # Перебираем все ячейки шапки между anchor_col и next_col (исключая разделители).
        col_to_metric: dict[int, str] = {}
        for c in range(anchor_col, next_col):
            label = cell_str(ws.cell(metric_row, c).value)
            if not label:
                continue
            col_to_metric[c] = label.lower()

        # Map field -> column.
        mapping: dict[str, int] = {}
        if metrics is not None:
            for field, aliases in metrics.items():
                aliases_low = tuple(a.lower() for a in aliases)
                for c, label_low in col_to_metric.items():
                    if any(skip in label_low for skip in skip_metric_substrings):
                        continue
                    if any(a in label_low for a in aliases_low):
                        mapping[field] = c
                        break
        else:
            aliases_low = tuple(a.lower() for a in metric_aliases)
            for c, label_low in col_to_metric.items():
                if any(a in label_low for a in aliases_low):
                    mapping[record_key or "value"] = c
                    break

        if mapping:
            fund_blocks.append((anchor_col, fund_name, mapping))

    if not fund_blocks:
        return []

    # 3. Идём по строкам данных, собираем записи (date, fund_name, metrics).
    # ``iter_data_rows`` ориентирован на колонку A; в wide-раскладке она часто пуста,
    # поэтому ходим напрямую по date_col и останавливаемся после серии пустых дат.
    out: list[dict] = []
    last_row = ws.max_row or 0
    empty_streak = 0
    for r in range(data_start, last_row + 1):
        d = cell_date(ws.cell(r, date_col).value)
        if d is None:
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0
        for _, fund_name, mapping in fund_blocks:
            record = {
                "snapshot_date": d,
                "cdu_name": fund_name,
                "src_row": r,
            }
            if defaults:
                record.update(defaults)
            any_value = False
            for field, col in mapping.items():
                v = cell_float(ws.cell(r, col).value)
                if v is not None:
                    any_value = True
                record[field] = v
            if any_value:
                out.append(record)
    return out


def _find_date_anchor(ws: Worksheet, *, max_scan: int = 30) -> tuple[Optional[int], Optional[int]]:
    """Найти ячейку со словом «Date»/«Дата» в первых ``max_scan`` строках."""
    upper = min(ws.max_row or 0, max_scan)
    last_col = ws.max_column or 0
    for r in range(1, upper + 1):
        for c in range(1, last_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                vl = v.strip().lower()
                if vl in ("date", "дата"):
                    return r, c
    return None, None


# ──────────────────────────────────────────────────────────────────────────
# Справочник — instrument reference
# ──────────────────────────────────────────────────────────────────────────
def parse_reference_sheet(ws: Worksheet) -> list[dict]:
    """Лист `Справочник` — каталог выпусков ЦБ."""
    header_row = find_header_row(ws, must_contain=("isin",))
    if header_row is None:
        return []

    headers = header_index_map(ws, header_row)
    col_isin = find_col(headers, "isin")
    col_ticker = find_col(headers, "ticker of kase", "ticker", "тикер")
    col_name = find_col(headers, "name", "наименование", "название")
    col_issuer = find_col(headers, "issuer", "эмитент")
    col_btype = find_col(headers, "bond type", "type", "тип")
    col_coupon = find_col(headers, "coupon", "купон")
    col_freq = find_col(headers, "frequency", "частота")
    col_base = find_col(headers, "base", "базис")
    col_nominal = find_col(headers, "nominal", "номинал")
    col_start = find_col(headers, "start date", "дата выпуска")
    col_maturity = find_col(headers, "maturity", "дата погашения")
    col_currency = find_col(headers, "currency", "валюта")

    if not col_isin:
        return []

    out: list[dict] = []
    for r in iter_data_rows(ws, header_row + 1):
        isin = cell_str(ws.cell(r, col_isin).value)
        if not isin or len(isin) < 6:
            continue
        out.append({
            "isin": isin,
            "ticker_kase": cell_str(ws.cell(r, col_ticker).value) if col_ticker else None,
            "instrument_name": cell_str(ws.cell(r, col_name).value) if col_name else None,
            "issuer": cell_str(ws.cell(r, col_issuer).value) if col_issuer else None,
            "bond_type": cell_str(ws.cell(r, col_btype).value) if col_btype else None,
            "coupon_rate_pct": cell_float(ws.cell(r, col_coupon).value) if col_coupon else None,
            "frequency": cell_int(ws.cell(r, col_freq).value) if col_freq else None,
            "base": cell_str(ws.cell(r, col_base).value) if col_base else None,
            "nominal": cell_float(ws.cell(r, col_nominal).value) if col_nominal else None,
            "start_date": cell_date(ws.cell(r, col_start).value) if col_start else None,
            "maturity_date": cell_date(ws.cell(r, col_maturity).value) if col_maturity else None,
            "currency": (cell_str(ws.cell(r, col_currency).value) or "KZT").upper()[:8] if col_currency else "KZT",
            "src_row": r,
        })
    return out


# ──────────────────────────────────────────────────────────────────────────
# FX rates — Нацбанк Казахстана, Доллар США
# ──────────────────────────────────────────────────────────────────────────
def parse_fx_sheet(ws: Worksheet) -> list[dict]:
    """Лист курсов USD/KZT от НБ РК. Ожидаемые колонки: Дата, Курс."""
    header_row = find_header_row(ws, must_contain=("дата",))
    if header_row is None:
        header_row = find_header_row(ws, must_contain=("курс",))
    if header_row is None:
        return []

    headers = header_index_map(ws, header_row)
    col_date = find_col(headers, "дата", "date")
    col_rate = find_col(headers, "курс", "rate", "официальный курс")
    if not col_date or not col_rate:
        return []

    out: list[dict] = []
    for r in iter_data_rows(ws, header_row + 1):
        d = cell_date(ws.cell(r, col_date).value)
        rate = cell_float(ws.cell(r, col_rate).value)
        if d is None or rate is None or rate <= 0:
            continue
        out.append({"rate_date": d, "currency": "USD", "rate": rate, "src_row": r})
    return out


# ──────────────────────────────────────────────────────────────────────────
# MBM Index history
# ──────────────────────────────────────────────────────────────────────────
def parse_mbm_sheet(ws: Worksheet) -> list[dict]:
    """Лист с историей MBM (KASE индекс)."""
    header_row = find_header_row(ws, must_contain=("дата",))
    if header_row is None:
        header_row = find_header_row(ws, must_contain=("date",))
    if header_row is None:
        return []

    headers = header_index_map(ws, header_row)
    col_date = find_col(headers, "дата", "date")
    col_ytm = find_col(headers, "ytm", "доходность", "yield")
    col_dur = find_col(headers, "duration", "дюрация")
    if not col_date:
        return []

    out: list[dict] = []
    for r in iter_data_rows(ws, header_row + 1):
        d = cell_date(ws.cell(r, col_date).value)
        if d is None:
            continue
        out.append({
            "index_date": d,
            "ytm_value": cell_float(ws.cell(r, col_ytm).value) if col_ytm else None,
            "duration": cell_float(ws.cell(r, col_dur).value) if col_dur else None,
        })
    return out


# ──────────────────────────────────────────────────────────────────────────
# Bond lots — ГЦБ / Агентские / МФО / Ин.ЦБ
# ──────────────────────────────────────────────────────────────────────────
def parse_bond_lots_sheet(ws: Worksheet, *, category: str,
                         fallback_date: Optional[date] = None) -> list[dict]:
    """Парсер любого из листов: ГЦБ / Агентские / МФО / Ин. ЦБ.

    Все они имеют схожую структуру: Sub portfolio name, ISIN, Trade date,
    Settlement date, Face value, Market price, Market value, YTM, Duration,
    Maturity, Last coupon date, Maturity status, etc.
    """
    header_row = find_header_row(ws, must_contain=("isin",))
    if header_row is None:
        return []

    headers = header_index_map(ws, header_row)
    col_isin = find_col(headers, "isin")
    col_sub = find_col(headers, "sub portfolio name", "sub-fund", "ду")
    col_trade = find_col(headers, "trade date", "дата сделки")
    col_settle = find_col(headers, "settlement date", "valuation date", "дата расчёта")
    col_face = find_col(headers, "face value", "номинал")
    col_qty = find_col(headers, "кол-во бумаг", "quantity", "кол-во")
    col_purch = find_col(headers, "purchase price", "цена покупки", "чистая цена")
    col_mprice = find_col(headers, "market price", "рыночная цена")
    col_acc = find_col(headers, "accrued interest", "нкд")
    col_mv = find_col(headers, "market value", "рыночная стоимость")
    col_total = find_col(headers, "total value")
    col_weight = find_col(headers, "weight", "вес")
    col_ytm = find_col(headers, "ytm", "доходность")
    col_dur = find_col(headers, "duration", "дюрация")
    col_maturity = find_col(headers, "maturity", "дата погашения")
    col_lastcoupon = find_col(headers, "last coupon date")
    col_status = find_col(headers, "maturity status", "статус")

    if not col_isin or not col_sub:
        return []

    out: list[dict] = []
    # В bond-листах колонка A пуста, ISIN живёт в col_isin → ходим по нему.
    for r in iter_data_rows(ws, header_row + 1, key_col=col_isin):
        isin = cell_str(ws.cell(r, col_isin).value)
        if not isin or len(isin) < 6:
            continue
        sub = normalize_subfund_name(ws.cell(r, col_sub).value)
        if not sub:
            continue
        face = cell_float(ws.cell(r, col_face).value) if col_face else None
        # Если лот списан в 0 — пропустим
        if face is not None and abs(face) < 0.01:
            continue
        out.append({
            "category": category,
            "isin": isin,
            "cdu_name": sub,
            # trade_date — дата исходной покупки/расчётов (из колонки Trade/Settlement);
            # valuation_date — дата СНИМКА портфеля (= report_date, fallback_date).
            "trade_date": (cell_date(ws.cell(r, col_trade).value) if col_trade else None)
                            or (cell_date(ws.cell(r, col_settle).value) if col_settle else None)
                            or fallback_date,
            "valuation_date": fallback_date,
            "face_value": face,
            "quantity": cell_float(ws.cell(r, col_qty).value) if col_qty else None,
            "purchase_price": cell_float(ws.cell(r, col_purch).value) if col_purch else None,
            "market_price": cell_float(ws.cell(r, col_mprice).value) if col_mprice else None,
            "accrued_interest": cell_float(ws.cell(r, col_acc).value) if col_acc else None,
            "market_value": cell_float(ws.cell(r, col_mv).value) if col_mv else None,
            "total_value": cell_float(ws.cell(r, col_total).value) if col_total else None,
            "weight": cell_float(ws.cell(r, col_weight).value) if col_weight else None,
            "ytm": cell_float(ws.cell(r, col_ytm).value) if col_ytm else None,
            "duration": cell_float(ws.cell(r, col_dur).value) if col_dur else None,
            "maturity_date": cell_date(ws.cell(r, col_maturity).value) if col_maturity else None,
            "last_coupon_date": cell_date(ws.cell(r, col_lastcoupon).value) if col_lastcoupon else None,
            "maturity_status": cell_str(ws.cell(r, col_status).value) if col_status else None,
            "src_row": r,
        })
    return out


# ──────────────────────────────────────────────────────────────────────────
# REPO open positions
# ──────────────────────────────────────────────────────────────────────────
def parse_repo_sheet(ws: Worksheet, *, fallback_date: Optional[date] = None) -> list[dict]:
    header_row = find_header_row(ws, must_contain=("close date",))
    if header_row is None:
        header_row = find_header_row(ws, must_contain=("face value",))
    if header_row is None:
        return []

    headers = header_index_map(ws, header_row)
    col_sub = find_col(headers, "sub portfolio name", "sub-fund")
    col_code = find_col(headers, "instrument code", "тикер", "ticker")
    col_isin = find_col(headers, "isin")
    col_trade = find_col(headers, "trade date")
    col_val = find_col(headers, "valuation date")
    col_close = find_col(headers, "close date")
    col_face = find_col(headers, "face value")
    col_closeval = find_col(headers, "close value")
    col_rate = find_col(headers, "repo rate", "ставка репо")
    col_acc = find_col(headers, "accrued interest")
    col_term = find_col(headers, "term", "срок")
    col_mv = find_col(headers, "market value")

    if not col_sub or not col_face:
        return []

    out: list[dict] = []
    for r in iter_data_rows(ws, header_row + 1):
        sub = normalize_subfund_name(ws.cell(r, col_sub).value)
        if not sub:
            continue
        face = cell_float(ws.cell(r, col_face).value) or 0.0
        if abs(face) < 0.01:
            continue
        out.append({
            "cdu_name": sub,
            "instrument_code": cell_str(ws.cell(r, col_code).value) if col_code else None,
            "isin": cell_str(ws.cell(r, col_isin).value) if col_isin else None,
            "trade_date": cell_date(ws.cell(r, col_trade).value) if col_trade else fallback_date,
            # valuation_date = report_date (дата снимка), не xlsx 'Valuation date' (= settle date).
            "valuation_date": fallback_date or (cell_date(ws.cell(r, col_val).value) if col_val else None),
            "close_date": cell_date(ws.cell(r, col_close).value) if col_close else None,
            "face_value": face,
            "close_value": cell_float(ws.cell(r, col_closeval).value) if col_closeval else None,
            "repo_rate_pct": cell_float(ws.cell(r, col_rate).value) if col_rate else None,
            "accrued_interest": cell_float(ws.cell(r, col_acc).value) if col_acc else None,
            "term_days": cell_int(ws.cell(r, col_term).value) if col_term else None,
            "market_value": cell_float(ws.cell(r, col_mv).value) if col_mv else None,
            "src_row": r,
        })
    return out


# ──────────────────────────────────────────────────────────────────────────
# Deposits — НБ РК
# ──────────────────────────────────────────────────────────────────────────
def parse_dep_sheet(ws: Worksheet, *, fallback_date: Optional[date] = None) -> list[dict]:
    header_row = find_header_row(ws, must_contain=("principal",))
    if header_row is None:
        header_row = find_header_row(ws, must_contain=("interest rate",))
    if header_row is None:
        return []

    headers = header_index_map(ws, header_row)
    col_sub = find_col(headers, "sub portfolio name", "sub-fund")
    col_trade = find_col(headers, "trade date")
    col_val = find_col(headers, "valuation date")
    col_close = find_col(headers, "close date")
    col_principal = find_col(headers, "principal", "номинал")
    col_rate = find_col(headers, "interest rate", "ставка")
    col_acc = find_col(headers, "accrued interest")
    col_mv = find_col(headers, "market value")

    if not col_sub or not col_principal:
        return []

    out: list[dict] = []
    for r in iter_data_rows(ws, header_row + 1):
        sub = normalize_subfund_name(ws.cell(r, col_sub).value)
        if not sub:
            continue
        principal = cell_float(ws.cell(r, col_principal).value) or 0.0
        if abs(principal) < 0.01:
            continue
        out.append({
            "cdu_name": sub,
            "trade_date": cell_date(ws.cell(r, col_trade).value) if col_trade else fallback_date,
            "valuation_date": fallback_date or (cell_date(ws.cell(r, col_val).value) if col_val else None),
            "close_date": cell_date(ws.cell(r, col_close).value) if col_close else None,
            "principal": principal,
            "interest_rate_pct": cell_float(ws.cell(r, col_rate).value) if col_rate else None,
            "accrued_interest": cell_float(ws.cell(r, col_acc).value) if col_acc else None,
            "market_value": cell_float(ws.cell(r, col_mv).value) if col_mv else None,
            "src_row": r,
        })
    return out


# ──────────────────────────────────────────────────────────────────────────
# Accounts receivable
# ──────────────────────────────────────────────────────────────────────────
def parse_ar_sheet(ws: Worksheet, *, fallback_date: Optional[date] = None) -> list[dict]:
    header_row = find_header_row(ws, must_contain=("isin",))
    if header_row is None:
        header_row = find_header_row(ws, must_contain=("остаток",))
    if header_row is None:
        return []

    headers = header_index_map(ws, header_row)
    col_du = find_col(headers, "ду", "sub-fund", "sub portfolio name")
    col_name = find_col(headers, "наименование", "description")
    col_isin = find_col(headers, "isin")
    col_ccy = find_col(headers, "валюта", "currency")
    col_record_date = find_col(headers, "дата постановки", "record date")
    col_balance_ccy = find_col(headers, "остаток в валюте", "balance currency")
    col_balance_kzt = find_col(headers, "остаток в тенге", "остаток в kzt", "balance kzt")
    col_due = find_col(headers, "дата завершения", "due date", "ожидаемая дата")

    if not col_isin or (not col_balance_kzt and not col_balance_ccy):
        return []

    out: list[dict] = []
    for r in iter_data_rows(ws, header_row + 1):
        isin = cell_str(ws.cell(r, col_isin).value)
        if not isin or len(isin) < 6:
            continue
        balance_kzt = cell_float(ws.cell(r, col_balance_kzt).value) if col_balance_kzt else 0.0
        balance_ccy = cell_float(ws.cell(r, col_balance_ccy).value) if col_balance_ccy else None
        if (balance_kzt is None or abs(balance_kzt) < 0.01) and \
           (balance_ccy is None or abs(balance_ccy or 0) < 0.01):
            continue
        sub = normalize_subfund_name(ws.cell(r, col_du).value) if col_du else None
        out.append({
            "cdu_name": sub,
            "isin": isin,
            "description": cell_str(ws.cell(r, col_name).value) if col_name else None,
            "currency": (cell_str(ws.cell(r, col_ccy).value) or "KZT").upper()[:8] if col_ccy else "KZT",
            "record_date": cell_date(ws.cell(r, col_record_date).value) if col_record_date else fallback_date,
            "balance_currency": balance_ccy,
            "balance_kzt": balance_kzt or 0.0,
            "due_date": cell_date(ws.cell(r, col_due).value) if col_due else None,
            "src_row": r,
        })
    return out


# ──────────────────────────────────────────────────────────────────────────
# Report sheet — извлечь Report!B2 (дату) и итоговые суммы по ЧДУ
# ──────────────────────────────────────────────────────────────────────────
def get_report_date(ws: Worksheet) -> Optional[date]:
    """Достать отчётную дату из Report!B2 (по соглашению из бизнес-процесса)."""
    if ws is None:
        return None
    return cell_date(ws.cell(2, 2).value)


def parse_report_sheet(ws: Worksheet, *, fallback_date: Optional[date] = None) -> list[dict]:
    """Parse the first Report sheet into dashboard summaries and category rows."""
    report_date = get_report_date(ws) or fallback_date
    if report_date is None:
        return []

    out: list[dict] = []
    current_cdu: Optional[str] = None
    pending_duration_cdu: Optional[str] = None

    def norm(value: Any) -> str:
        return (cell_str(value) or "").strip().lower()

    def row_contains(row: int, needle: str) -> bool:
        needle_l = needle.lower()
        for c in range(1, (ws.max_column or 0) + 1):
            if needle_l in norm(ws.cell(row, c).value):
                return True
        return False

    def col_by_alias(row: int, *aliases: str) -> Optional[int]:
        aliases_l = tuple(a.lower() for a in aliases)
        for c in range(1, (ws.max_column or 0) + 1):
            label = norm(ws.cell(row, c).value)
            if label and any(a == label or a in label for a in aliases_l):
                return c
        return None

    for r in range(1, (ws.max_row or 0) + 1):
        maybe_cdu = normalize_subfund_name(ws.cell(r, 2).value)
        if maybe_cdu and not row_contains(r, "instruments"):
            current_cdu = maybe_cdu

        if pending_duration_cdu and norm(ws.cell(r, 3).value) not in ("", "duration"):
            summary = _last_report_summary(out, pending_duration_cdu)
            if summary:
                summary["benchmark_duration"] = _report_duration_scale(cell_float(ws.cell(r, 4).value))
                summary["duration_lower"] = _report_duration_scale(cell_float(ws.cell(r, 5).value))
                summary["duration_upper"] = _report_duration_scale(cell_float(ws.cell(r, 6).value))
                summary["duration_status"] = cell_str(ws.cell(r, 7).value)
            pending_duration_cdu = None

        if norm(ws.cell(r, 3).value) == "duration":
            pending_duration_cdu = current_cdu
            continue

        instr_col = col_by_alias(r, "instruments")
        if instr_col is None or not current_cdu:
            continue

        col_prev = col_by_alias(r, "market value t-1")
        col_change = col_by_alias(r, "daily change")
        col_current = col_by_alias(r, "current market value")
        col_pct = col_by_alias(r, "% of total invest")
        col_ytm = col_by_alias(r, "ytm")
        col_duration = col_by_alias(r, "duration")
        col_min = col_by_alias(r, "min. limit", "min limit")
        col_max = col_by_alias(r, "max limit")
        col_hard = col_by_alias(r, "hard limit")
        col_soft = col_by_alias(r, "soft limit")
        col_free = col_by_alias(r, "свободный остаток", "free limit")

        empty_streak = 0
        block_positions = 0
        for data_row in range(r + 1, (ws.max_row or 0) + 1):
            instrument = cell_str(ws.cell(data_row, instr_col).value)
            if not instrument:
                current_mv = cell_float(ws.cell(data_row, col_current).value) if col_current else None
                pct_total = cell_float(ws.cell(data_row, col_pct).value) if col_pct else None
                if block_positions and current_mv is not None and (
                    pct_total is None or abs(pct_total - 1.0) < 0.0001
                ):
                    out.append({
                        "kind": "summary",
                        "snapshot_date": report_date,
                        "cdu_name": current_cdu,
                        "total_mv_prev": (cell_float(ws.cell(data_row, col_prev).value) if col_prev else None) or 0.0,
                        "total_daily_change": (cell_float(ws.cell(data_row, col_change).value) if col_change else None) or 0.0,
                        "total_mv_current": current_mv or 0.0,
                        "ytm_weighted": (cell_float(ws.cell(data_row, col_ytm).value) if col_ytm else None) or 0.0,
                        "duration_weighted": (cell_float(ws.cell(data_row, col_duration).value) if col_duration else None) or 0.0,
                        "src_row": data_row,
                    })
                    break
                empty_streak += 1
                if empty_streak >= 2:
                    break
                continue
            empty_streak = 0
            if normalize_subfund_name(instrument):
                break

            current_mv = cell_float(ws.cell(data_row, col_current).value) if col_current else None
            if instrument.strip().lower().startswith("total"):
                out.append({
                    "kind": "summary",
                    "snapshot_date": report_date,
                    "cdu_name": current_cdu,
                    "total_mv_prev": (cell_float(ws.cell(data_row, col_prev).value) if col_prev else None) or 0.0,
                    "total_daily_change": (cell_float(ws.cell(data_row, col_change).value) if col_change else None) or 0.0,
                    "total_mv_current": current_mv or 0.0,
                    "ytm_weighted": (cell_float(ws.cell(data_row, col_ytm).value) if col_ytm else None) or 0.0,
                    "duration_weighted": (cell_float(ws.cell(data_row, col_duration).value) if col_duration else None) or 0.0,
                    "src_row": data_row,
                })
                break

            category = _report_category(instrument)
            if category is None:
                continue
            out.append({
                "kind": "position",
                "snapshot_date": report_date,
                "cdu_name": current_cdu,
                "category": category,
                "instrument_name": instrument,
                "market_value_prev": (cell_float(ws.cell(data_row, col_prev).value) if col_prev else None) or 0.0,
                "daily_change": (cell_float(ws.cell(data_row, col_change).value) if col_change else None) or 0.0,
                "market_value_current": current_mv or 0.0,
                "pct_of_total": (cell_float(ws.cell(data_row, col_pct).value) if col_pct else None) or 0.0,
                "ytm": cell_float(ws.cell(data_row, col_ytm).value) if col_ytm else None,
                "duration": cell_float(ws.cell(data_row, col_duration).value) if col_duration else None,
                "min_limit_pct": cell_float(ws.cell(data_row, col_min).value) if col_min else None,
                "max_limit_pct": cell_float(ws.cell(data_row, col_max).value) if col_max else None,
                "hard_limit": cell_str(ws.cell(data_row, col_hard).value) if col_hard else None,
                "soft_limit": cell_str(ws.cell(data_row, col_soft).value) if col_soft else None,
                "free_limit_mln": cell_float(ws.cell(data_row, col_free).value) if col_free else None,
                "src_row": data_row,
            })
            block_positions += 1

    return out


def _report_category(label: str) -> Optional[str]:
    value = label.strip().lower()
    if "cash" in value:
        return "CASH"
    if "государ" in value or "гцб" in value:
        return "GOV_BONDS"
    if "repo" in value or "репо" in value:
        return "REVERSE_REPO"
    if "мфо" in value:
        return "MFO_BONDS"
    if "агент" in value:
        return "AGENCY_BONDS"
    if "иностран" in value:
        return "FOREIGN_BONDS"
    if "депозит" in value:
        return "DEPOSIT"
    if "дебитор" in value or "receivable" in value:
        return "RECEIVABLES"
    return None


def _last_report_summary(rows: list[dict], cdu_name: str) -> Optional[dict]:
    for row in reversed(rows):
        if row.get("kind") == "summary" and row.get("cdu_name") == cdu_name:
            return row
    return None


def _report_duration_scale(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    return value * 100.0 if abs(value) < 0.5 else value
