"""Parse Holdings report XLSX (from ЧДУ custodians) into CashSnapshot + PortfolioPosition.

Expected layout (no header row; data starts immediately):
  Col 0: counterparty_name (e.g. 'АО «Tansar Capital»')
  Col 1: custodian_name    (e.g. 'АО "Банк ЦентрКредит"')
  Col 2: internal_id       (e.g. 'I+1CACE30098')
  Col 3: report_date       (datetime)
  Col 4: instrument_type   ('bond', 'repo', None for cash rows)
  Col 5: ISIN / cash label ('KZKD...', 'XS...', 'cash KZT', 'cash USD')
  Col 6: quantity / amount (numeric)
  Col 7+: None
"""
from __future__ import annotations
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import hashlib
from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session
from models.db_models import CDU, CashSnapshot, PortfolioPosition, SourceDocument
from services.calculator.constants import normalize_cdu_name
from services.parser.classification import detect_cdu_prefix, DEFAULT_CDU_PREFIXES


def parse_holdings_xlsx(file_path: Path | str) -> Dict[str, Any]:
    wb = load_workbook(Path(file_path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"positions": [], "cash": [], "warnings": ["Empty sheet"]}

    # Try to detect CDU from filename
    filename = Path(file_path).name
    cdu_name = None
    for prefix, name in sorted(DEFAULT_CDU_PREFIXES.items(), key=lambda x: -len(x[0])):
        if prefix.lower() in filename.lower():
            cdu_name = name
            break

    report_date: Optional[date] = None
    positions: List[Dict] = []
    cash_rows: List[Dict] = []
    warnings: List[str] = []

    for i, row in enumerate(rows):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        # Skip description/legend rows (long text in col4 or col5)
        if row[4] and isinstance(row[4], str) and len(row[4]) > 50:
            continue
        if row[5] and isinstance(row[5], str) and len(row[5]) > 50:
            continue

        # Extract date from first valid row if not yet set
        if report_date is None and len(row) > 3 and isinstance(row[3], datetime):
            report_date = row[3].date()
        if report_date is None and len(row) > 3 and isinstance(row[3], date):
            report_date = row[3]

        inst_type = str(row[4]).strip().lower() if len(row) > 4 and row[4] else None
        label = str(row[5]).strip() if len(row) > 5 and row[5] else None
        qty = _to_float(row[6]) if len(row) > 6 else None

        if not label:
            continue

        # Skip header/legend rows (ISIN header, empty quantity)
        if label.upper() in ("ISIN", "INSTRUMENT", "ИНСТРУМЕНТ", "КОЛИЧЕСТВО ЦБ") or qty is None:
            continue

        # Detect CDU from row if not found from filename
        if cdu_name is None and row[0]:
            raw = str(row[0])
            for prefix, name in sorted(DEFAULT_CDU_PREFIXES.items(), key=lambda x: -len(x[0])):
                if prefix.lower() in raw.lower():
                    cdu_name = name
                    break

        if label.lower().startswith("cash"):
            ccy = "KZT"
            if "usd" in label.lower():
                ccy = "USD"
            cash_rows.append({"currency": ccy, "amount": qty or 0.0})
        else:
            isin = label
            category = "OTHER"
            if inst_type == "bond":
                category = _guess_bond_category(isin)
            elif inst_type == "repo":
                category = "REVERSE_REPO"
            positions.append({
                "isin": isin,
                "instrument_type": inst_type,
                "category": category,
                "quantity": qty or 0.0,
                "row_idx": i,
            })

    if cdu_name is None:
        warnings.append("CDU not detected from filename or content")
    return {
        "cdu_name": cdu_name,
        "report_date": report_date,
        "positions": positions,
        "cash": cash_rows,
        "warnings": warnings,
    }


def import_holdings_xlsx(
    db: Session,
    file_path: Path | str,
    *,
    uploaded_by: Optional[str] = None,
    source_doc_id: Optional[int] = None,
) -> Dict[str, int]:
    parsed = parse_holdings_xlsx(file_path)
    cdu_name = parsed["cdu_name"]
    report_date = parsed["report_date"] or date.today()

    cdu_id: Optional[int] = None
    if cdu_name:
        canonical = normalize_cdu_name(cdu_name)
        if canonical:
            row = db.execute(select(CDU).where(CDU.name == canonical)).scalars().first()
            if row:
                cdu_id = row.id

    if cdu_id is None:
        logger.warning(f"Holdings import: CDU not resolved for {file_path}")
        return {"cash_snapshots": 0, "portfolio_positions": 0, "skipped": 1}

    counters = {"cash_snapshots": 0, "portfolio_positions": 0, "skipped": 0}

    # Upsert cash snapshots
    for cr in parsed["cash"]:
        ccy = cr["currency"]
        amt = cr["amount"]
        existing = db.execute(select(CashSnapshot).where(
            CashSnapshot.cdu_id == cdu_id,
            CashSnapshot.snapshot_date == report_date,
            CashSnapshot.currency == ccy,
        )).scalars().first()
        if existing:
            existing.amount = amt
            existing.source = "holdings_report_xlsx"
            existing.source_doc_id = source_doc_id
        else:
            db.add(CashSnapshot(
                cdu_id=cdu_id, snapshot_date=report_date,
                currency=ccy, amount=amt,
                source="holdings_report_xlsx", source_doc_id=source_doc_id,
            ))
        counters["cash_snapshots"] += 1

    # Upsert portfolio positions
    for pos in parsed["positions"]:
        isin = pos["isin"]
        qty = pos["quantity"]
        existing = db.execute(select(PortfolioPosition).where(
            PortfolioPosition.cdu_id == cdu_id,
            PortfolioPosition.position_date == report_date,
            PortfolioPosition.instrument_code == isin,
        )).scalars().first()
        if existing:
            existing.nominal_volume = qty
            existing.instrument_category = pos["category"]
            existing.instrument_name = pos.get("instrument_type")
        else:
            db.add(PortfolioPosition(
                cdu_id=cdu_id, position_date=report_date,
                instrument_code=isin, nominal_volume=qty,
                instrument_category=pos["category"],
                instrument_name=pos.get("instrument_type"),
            ))
        counters["portfolio_positions"] += 1

    db.commit()
    logger.info(f"Holdings import: {counters} cdu={cdu_id} date={report_date}")
    return counters


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return None if v != v else float(v)
    try:
        s = str(v).replace(" ", "").replace("\u00a0", "").replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return None


def _guess_bond_category(isin: str) -> str:
    if isin.startswith(("KFUS", "MFRK")):
        return "GOV_BONDS"
    if isin.startswith(("EABR", "EAB")):
        return "AGENCY_BONDS"
    if isin.startswith("MFO"):
        return "MFO_BONDS"
    if isin.startswith(("XS", "US", "RU")):
        return "FOREIGN_BONDS"
    return "OTHER"
