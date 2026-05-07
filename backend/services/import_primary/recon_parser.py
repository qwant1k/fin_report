"""Reconciliation XLSX parser.

Expected layout (from ЧДУ):
  - Securities table with an "ИТОГО" row.
  - Cash row "Остатки на инвест. счетах".
  - REPO table "Обратное репо" with current-value column.
  - AR block and "Итого активов".
"""
from __future__ import annotations
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook


def parse_reconciliation_xlsx(file_path: Path | str) -> Dict[str, Any]:
    wb = load_workbook(Path(file_path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    totals: Dict[str, float] = {}
    warnings: List[str] = []
    report_date: Optional[date] = None
    cdu_name: Optional[str] = None
    section: Optional[str] = None
    repo_sum = 0.0
    ar_sum = 0.0

    for row in rows:
        cells = list(row or ())
        if not cells:
            continue
        text_cells = [str(c).strip() for c in cells if isinstance(c, str) and str(c).strip()]
        row_text = " | ".join(text_cells).lower()

        if cdu_name is None and any("доверительного управляющего" in t.lower() for t in text_cells):
            for c in cells:
                if isinstance(c, str) and ("АО" in c or "Invest" in c or "Finance" in c):
                    cdu_name = c.strip()
                    break
        if report_date is None:
            for c in cells:
                report_date = _to_date(c)
                if report_date:
                    break

        if "цб" == row_text or row_text.endswith("цб") or "ценные бумаги" in row_text:
            section = "securities"
            continue
        if "обратное репо" in row_text or "repo" in row_text:
            section = "repo"
            continue
        if "прочие активы" in row_text or "дебитор" in row_text:
            section = "ar"
        if "кредиторская задолженность" in row_text:
            section = None

        nums = [_to_float(c) for c in cells]
        nums_present = [n for n in nums if n is not None]

        if any("остатки на инвест" in t.lower() or "cash" in t.lower() for t in text_cells):
            amount = _first_numeric_after_text(cells)
            if amount is not None:
                totals["cash"] = totals.get("cash", 0.0) + amount

        if section == "securities" and any(t.strip().upper() == "ИТОГО" for t in text_cells):
            # In real ЧДУ sheets the current market value is usually the last
            # large numeric before helper/check columns.
            if nums_present:
                totals["securities"] = _best_total_value(nums_present)

        if section == "repo" and cells and _to_float(cells[0]) is not None:
            repo_value = _last_numeric(cells)
            if repo_value is not None:
                repo_sum += repo_value

        if section == "ar" and cells and _to_float(cells[0]) is not None:
            ar_value = _last_numeric(cells)
            if ar_value is not None:
                ar_sum += ar_value

        if any("итого активов" in t.lower() or "total assets" in t.lower() for t in text_cells):
            amount = _first_numeric_after_text(cells)
            if amount is not None:
                totals["total"] = amount

    if repo_sum:
        totals["repo"] = repo_sum
    totals.setdefault("ar", ar_sum)
    for key in ("cash", "securities", "repo", "ar", "total"):
        if key not in totals:
            warnings.append(f"Не найден итог сверки: {key}")
    return {
        "cdu_name": cdu_name,
        "report_date": report_date,
        "totals": totals,
        "rows_parsed": len(rows),
        "warnings": warnings,
    }


def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return None if v != v else float(v)
    try:
        s = str(v).replace(" ", "").replace("\u00a0", "").replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return None


def _to_date(v: Any) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)) and 1 <= float(v) <= 60000:
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
    if isinstance(v, str):
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _first_numeric_after_text(cells: List[Any]) -> Optional[float]:
    seen_text = False
    for c in cells:
        if isinstance(c, str) and c.strip():
            seen_text = True
            continue
        if seen_text:
            n = _to_float(c)
            if n is not None:
                return n
    return None


def _last_numeric(cells: List[Any]) -> Optional[float]:
    for c in reversed(cells):
        n = _to_float(c)
        if n is not None:
            return n
    return None


def _best_total_value(nums: List[float]) -> float:
    # Ignore percentages/check columns where possible and prefer the last
    # large money amount in the row.
    large = [n for n in nums if abs(n) > 1000]
    return large[-1] if large else nums[-1]
