"""Generate the consolidated risk report XLSX exactly like the manual file.

Layout per CDU block (mirrors the original `Report` sheet):
┌────────────────────────────────────────────────────────────────────┐
│  <CDU name>                                                       │  ← header #1F6B38
├──────────────────────────────────────────────────────────────────┤
│ Instruments | MV T-1 | Daily change | CMV | %Total | YTM | Dur. | Min | Max | Hard | Soft | Free (mln) │
├──────────────────────────────────────────────────────────────────┤
│  Cash                                                             │
│  ГЦБ                                                              │
│  Обратное REPO                              ← yellow #FFFF00     │
│  МФО                                                              │
│  Агентские                                                        │
│  Дебиторка                                                        │
├──────────────────────────────────────────────────────────────────┤
│  Total: …                                  ← green  #70AD47       │
└──────────────────────────────────────────────────────────────────┘
   <CDU share %>
   Duration | benchmark | -0.2 | +0.5 | limit ok/breach
"""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import select
from sqlalchemy.orm import Session

from models.db_models import (
    Alert,
    CDU,
    KasePrice,
    MBMIndex,
    PortfolioPosition,
    PortfolioSummary,
)
from services.calculator.constants import (
    CATEGORY_LABELS,
    CATEGORY_ORDER,
    DEFAULT_LIMITS,
    DURATION_LOWER_OFFSET,
    DURATION_UPPER_OFFSET,
)


# ──────── styles ────────
HEADER_FILL = PatternFill("solid", fgColor="1F6B38")
COLUMN_FILL = PatternFill("solid", fgColor="4CAF50")
REPO_FILL = PatternFill("solid", fgColor="FFFF00")
TOTAL_FILL = PatternFill("solid", fgColor="70AD47")
ERROR_FILL = PatternFill("solid", fgColor="FF0000")
ALT_FILL = PatternFill("solid", fgColor="F4F8F4")

WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BLACK_BOLD = Font(name="Calibri", size=11, bold=True, color="000000")
BLACK = Font(name="Calibri", size=11, color="000000")
WHITE = Font(name="Calibri", size=11, color="FFFFFF")

THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT = '# ##0.00;-# ##0.00;"-"'
PCT_FMT = '0.00%'
INT_FMT = '# ##0'


COLUMNS = [
    ("Instruments", 36),
    ("Market Value T-1", 18),
    ("Daily change", 16),
    ("Current Market Value", 22),
    ("% of Total Invest", 14),
    ("YTM", 10),
    ("Duration", 10),
    ("Min. Limit", 11),
    ("Max Limit", 11),
    ("Hard limit", 11),
    ("Soft Limit", 11),
    ("Свободный остаток (млн ₸)", 22),
]


def generate_xlsx_report(db: Session, report_date: date, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    out = output_dir / f"risk_report_{report_date.strftime('%Y%m%d')}.xlsx"

    wb = Workbook()
    _write_main_sheet(wb, db, report_date)
    _write_kase_sheet(wb, db, report_date)
    _write_history_sheet(wb, db, report_date)
    _write_alerts_sheet(wb, db, report_date)

    wb.save(out)
    return out


# ─────── main sheet ───────
def _write_main_sheet(wb: Workbook, db: Session, report_date: date) -> None:
    ws = wb.active
    ws.title = "Сводный отчёт"

    # report-level header
    ws["A1"] = "Date"
    ws["A1"].font = BLACK_BOLD
    ws["B1"] = report_date
    ws["B1"].number_format = "DD.MM.YYYY"

    summaries = db.execute(select(PortfolioSummary).where(
        PortfolioSummary.summary_date == report_date,
    )).scalars().all()
    if not summaries:
        ws["A3"] = "Нет данных за выбранную дату."
        return

    fund_total = sum(s.total_mv_current for s in summaries)
    fund_total_prev = sum(s.total_mv_prev for s in summaries)
    fund_change = fund_total - fund_total_prev

    ws["A2"] = "Все активы Фонда"
    ws["A2"].font = BLACK_BOLD
    ws["B2"] = fund_total
    ws["B2"].number_format = NUM_FMT
    ws["C2"] = "Изменение за день"
    ws["D2"] = fund_change
    ws["D2"].number_format = NUM_FMT
    if fund_change < 0:
        ws["D2"].font = Font(color="C00000", bold=True)
    elif fund_change > 0:
        ws["D2"].font = Font(color="2E7D32", bold=True)

    row = 5
    for s in summaries:
        cdu = db.get(CDU, s.cdu_id)
        if not cdu:
            continue
        positions = db.execute(select(PortfolioPosition).where(
            PortfolioPosition.cdu_id == s.cdu_id,
            PortfolioPosition.position_date == report_date,
        )).scalars().all()
        positions_by_cat = {p.instrument_category: p for p in positions}

        row = _write_cdu_block(ws, row, cdu, s, positions_by_cat, fund_total)
        row += 2  # blank rows between blocks

    # column widths
    for i, (_, width) in enumerate(COLUMNS, start=2):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.column_dimensions["A"].width = 20

    ws.freeze_panes = "C5"


def _write_cdu_block(ws, start_row, cdu, summary, pos_by_cat, fund_total):
    # CDU title row
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row,
                   end_column=1 + len(COLUMNS))
    cell = ws.cell(start_row, 1, cdu.name)
    cell.fill = HEADER_FILL
    cell.font = WHITE_BOLD
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[start_row].height = 24

    # column headers
    header_row = start_row + 1
    ws.cell(header_row, 1, "").fill = COLUMN_FILL
    for i, (label, _w) in enumerate(COLUMNS, start=2):
        c = ws.cell(header_row, i, label)
        c.fill = COLUMN_FILL
        c.font = WHITE_BOLD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[header_row].height = 30

    # data rows
    r = header_row + 1
    for cat in CATEGORY_ORDER:
        pos = pos_by_cat.get(cat)
        mn, mx = DEFAULT_LIMITS.get(cat, (0.0, 1.0))
        row_data = {
            "label": CATEGORY_LABELS[cat],
            "mv_prev": pos.market_value_prev if pos else 0.0,
            "daily": pos.daily_change if pos else 0.0,
            "cmv": pos.market_value_current if pos else 0.0,
            "pct": pos.pct_of_total if pos else 0.0,
            "ytm": (pos.ytm or 0.0) if pos else 0.0,
            "dur": (pos.duration or 0.0) if pos else 0.0,
            "min": mn,
            "max": mx,
            "hard": (pos.hard_limit_status or "ok") if pos else "ok",
            "soft": (pos.soft_limit_status or "ok") if pos else "ok",
            "free": (pos.free_limit_mln if pos and pos.free_limit_mln is not None else None),
        }
        is_repo = cat == "REVERSE_REPO"
        is_breach = row_data["hard"] == "breach"
        _write_data_row(ws, r, row_data, repo=is_repo, breach=is_breach)
        r += 1

    # Total row
    total_row = r
    total_data = {
        "label": "Total:",
        "mv_prev": summary.total_mv_prev,
        "daily": summary.total_daily_change,
        "cmv": summary.total_mv_current,
        "pct": 1.0,
        "ytm": summary.ytm_weighted,
        "dur": summary.duration_weighted,
        "min": "",
        "max": "",
        "hard": "",
        "soft": "",
        "free": None,
    }
    _write_data_row(ws, total_row, total_data, total=True)
    r += 1

    # CDU-share row
    ws.cell(r, 1, "Доля в Фонде").font = BLACK_BOLD
    cell = ws.cell(r, 2, summary.cdu_share_pct)
    cell.number_format = PCT_FMT
    cell.font = BLACK_BOLD
    r += 1

    # Duration row
    ws.cell(r, 1, "Duration").font = BLACK_BOLD
    ws.cell(r, 2, summary.duration_weighted).number_format = NUM_FMT
    ws.cell(r, 3, "benchmark duration").font = BLACK
    ws.cell(r, 4, summary.benchmark_duration if summary.benchmark_duration is not None else "—")
    if summary.benchmark_duration is not None:
        ws.cell(r, 4).number_format = NUM_FMT
        ws.cell(r, 5, summary.benchmark_duration + DURATION_LOWER_OFFSET).number_format = NUM_FMT
        ws.cell(r, 6, summary.benchmark_duration + DURATION_UPPER_OFFSET).number_format = NUM_FMT
    ws.cell(r, 7, summary.duration_status or "—")
    if summary.duration_status == "breach":
        ws.cell(r, 7).fill = ERROR_FILL
        ws.cell(r, 7).font = WHITE_BOLD
    elif summary.duration_status == "ok":
        ws.cell(r, 7).fill = TOTAL_FILL
        ws.cell(r, 7).font = WHITE_BOLD
    return r + 1


def _write_data_row(ws, row, data, *, repo: bool = False, breach: bool = False, total: bool = False) -> None:
    fill = None
    font = BLACK
    if total:
        fill = TOTAL_FILL
        font = WHITE_BOLD
    elif breach:
        fill = ERROR_FILL
        font = WHITE_BOLD
    elif repo:
        fill = REPO_FILL
        font = BLACK_BOLD

    cells = [
        (1, data["label"], None),
        (2, data["mv_prev"], NUM_FMT),
        (3, data["daily"], NUM_FMT),
        (4, data["cmv"], NUM_FMT),
        (5, data["pct"], PCT_FMT),
        (6, data["ytm"], PCT_FMT if isinstance(data["ytm"], float) and abs(data["ytm"]) < 1 else NUM_FMT),
        (7, data["dur"], NUM_FMT),
        (8, data["min"], PCT_FMT if isinstance(data["min"], float) else None),
        (9, data["max"], PCT_FMT if isinstance(data["max"], float) else None),
        (10, data["hard"], None),
        (11, data["soft"], None),
        (12, (data["free"] if data["free"] is not None else "—"), NUM_FMT if isinstance(data["free"], float) else None),
    ]
    for col, val, fmt in cells:
        c = ws.cell(row, col, val)
        c.border = BORDER
        if fill is not None:
            c.fill = fill
        c.font = font
        if fmt:
            c.number_format = fmt
        if col == 1:
            c.alignment = Alignment(horizontal="left", vertical="center")
        else:
            c.alignment = Alignment(horizontal="right", vertical="center")
        # paint hard/soft red individually if breach
        if not total and col in (10, 11) and val == "breach":
            c.fill = ERROR_FILL
            c.font = WHITE_BOLD


# ─────── KASE sheet ───────
def _write_kase_sheet(wb: Workbook, db: Session, report_date: date) -> None:
    ws = wb.create_sheet("Сверка KASE")
    headers = ["Инструмент", "ISIN", "Наименование", "Close", "YTM", "НКД", "Duration", "Источник"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(1, i, h)
        c.fill = COLUMN_FILL
        c.font = WHITE_BOLD
        c.border = BORDER
    rows = db.execute(select(KasePrice).where(KasePrice.trade_date == report_date)).scalars().all()
    for i, r in enumerate(rows, start=2):
        ws.cell(i, 1, r.instrument_code)
        ws.cell(i, 2, r.isin or "")
        ws.cell(i, 3, r.instrument_name or "")
        ws.cell(i, 4, r.close_price).number_format = NUM_FMT
        ws.cell(i, 5, r.ytm).number_format = NUM_FMT
        ws.cell(i, 6, r.accrued_interest).number_format = NUM_FMT
        ws.cell(i, 7, r.duration).number_format = NUM_FMT
        ws.cell(i, 8, r.source)
    for i, w in enumerate([18, 16, 36, 14, 10, 14, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────── History sheet ───────
def _write_history_sheet(wb: Workbook, db: Session, report_date: date) -> None:
    ws = wb.create_sheet("История")
    headers = ["Дата", "ЧДУ", "MV Total", "Daily Change", "YTM", "Duration"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(1, i, h)
        c.fill = COLUMN_FILL
        c.font = WHITE_BOLD
    rows = db.execute(select(PortfolioSummary).order_by(PortfolioSummary.summary_date.desc()).limit(500)).scalars().all()
    for i, s in enumerate(rows, start=2):
        cdu = db.get(CDU, s.cdu_id)
        ws.cell(i, 1, s.summary_date).number_format = "DD.MM.YYYY"
        ws.cell(i, 2, cdu.short_name if cdu else "")
        ws.cell(i, 3, s.total_mv_current).number_format = NUM_FMT
        ws.cell(i, 4, s.total_daily_change).number_format = NUM_FMT
        ws.cell(i, 5, s.ytm_weighted).number_format = PCT_FMT
        ws.cell(i, 6, s.duration_weighted).number_format = NUM_FMT
    for i, w in enumerate([12, 18, 18, 16, 10, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────── Alerts sheet ───────
def _write_alerts_sheet(wb: Workbook, db: Session, report_date: date) -> None:
    ws = wb.create_sheet("Алерты")
    headers = ["Дата", "ЧДУ", "Тип", "Severity", "Сообщение", "Решено"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(1, i, h)
        c.fill = COLUMN_FILL
        c.font = WHITE_BOLD
    rows = db.execute(select(Alert).where(Alert.alert_date == report_date)).scalars().all()
    for i, a in enumerate(rows, start=2):
        cdu = db.get(CDU, a.cdu_id) if a.cdu_id else None
        ws.cell(i, 1, a.alert_date).number_format = "DD.MM.YYYY"
        ws.cell(i, 2, cdu.short_name if cdu else "")
        ws.cell(i, 3, a.alert_type)
        ws.cell(i, 4, a.severity)
        ws.cell(i, 5, a.message)
        ws.cell(i, 6, "да" if a.is_resolved else "нет")
        if a.severity == "CRITICAL":
            ws.cell(i, 4).fill = ERROR_FILL
            ws.cell(i, 4).font = WHITE_BOLD
    for i, w in enumerate([12, 18, 18, 12, 60, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
