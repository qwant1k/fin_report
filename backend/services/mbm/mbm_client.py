"""MBM (Money Benchmark) — fetch daily MBM index / Duration / ModDuration.

Источники, в порядке приоритета:

1. **KASE XLSX archive** — официальный экспорт со страницы
   https://kase.kz/ru/indexes-and-indicators/composite-indexes/mbm-index .
   Эндпоинт ``/api/indicators/mbm-index/archive-xls`` принимает диапазон
   ``start_date``/``end_date`` (YYYY-MM-DD) и отдаёт XLSX с колонками::

       Дата расчетов | MBM index | Duration | ModDuration

   Числа даны в "американском" формате (``1,214.9000`` = 1 214.9), даты —
   ``ДД.ММ.ГГГГ`` строкой. Это самый надёжный источник: страница MBM —
   SPA-приложение и не отдаёт значения в HTML.

2. **NBRK HTML** — резервный, как раньше (``settings.nbrk_mbm_url``).
3. **Ручной ввод** через ``/api/mbm/manual``.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import List, Optional

import httpx
from bs4 import BeautifulSoup
from loguru import logger
from openpyxl import load_workbook

from config import settings


@dataclass
class MBMValue:
    index_date: date
    ytm_value: Optional[float]   # колонка "MBM index" из XLSX (уровень индекса)
    duration: Optional[float]
    mod_duration: Optional[float] = None
    source: str = ""


class MBMClient:
    """Загрузчик MBM-индекса с KASE/НБ РК."""

    def __init__(self) -> None:
        self.nbrk_url = settings.nbrk_mbm_url
        self.kase_xlsx_url = settings.kase_mbm_xlsx_url
        self.lookback_days = max(1, int(getattr(settings, "kase_mbm_lookback_days", 14)))
        self._headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            ),
            "Accept-Language": "ru,en;q=0.8",
        }

    # ────────────────── public API ──────────────────
    async def fetch_latest(self) -> Optional[MBMValue]:
        """Вернуть самое свежее доступное значение MBM (за последние N дней)."""
        rows = await self.fetch_history()
        return rows[0] if rows else None

    async def fetch_for_date(self, target: date) -> Optional[MBMValue]:
        """Получить значение строго на указанную дату (или None, если KASE его не отдаёт)."""
        rows = await self.fetch_history(start=target, end=target)
        for r in rows:
            if r.index_date == target:
                return r
        return None

    async def fetch_history(
        self,
        start: Optional[date] = None,
        end: Optional[date] = None,
    ) -> List[MBMValue]:
        """Скачать диапазон значений MBM, отсортированных по убыванию даты."""
        end = end or date.today()
        start = start or (end - timedelta(days=self.lookback_days))
        async with httpx.AsyncClient(
            timeout=30.0, headers=self._headers, follow_redirects=True
        ) as cli:
            # 1) KASE XLSX архив — основной источник.
            try:
                rows = await self._fetch_kase_xlsx(cli, start, end)
                if rows:
                    return rows
            except Exception as exc:  # pragma: no cover - сетевые сбои
                logger.warning(f"MBM KASE XLSX fetch failed: {exc!r}")

            # 2) NBRK HTML — резерв.
            try:
                v = await self._fetch_nbrk(cli)
                if v:
                    return [v]
            except Exception as exc:  # pragma: no cover
                logger.warning(f"MBM NBRK fetch failed: {exc!r}")
        return []

    # ────────────────── KASE XLSX ──────────────────
    async def _fetch_kase_xlsx(
        self, cli: httpx.AsyncClient, start: date, end: date,
    ) -> List[MBMValue]:
        params = {
            "start_date": start.strftime("%Y-%m-%d"),
            "end_date": end.strftime("%Y-%m-%d"),
            "language": "ru",
        }
        r = await cli.get(self.kase_xlsx_url, params=params)
        r.raise_for_status()
        ctype = r.headers.get("Content-Type", "")
        if "spreadsheet" not in ctype and not r.content[:2] == b"PK":
            logger.warning(
                f"MBM KASE XLSX: unexpected Content-Type={ctype!r}, len={len(r.content)}"
            )
            return []
        return parse_kase_mbm_xlsx(r.content)

    # ────────────────── NBRK fallback ──────────────────
    async def _fetch_nbrk(self, cli: httpx.AsyncClient) -> Optional[MBMValue]:
        r = await cli.get(self.nbrk_url)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")
        text = soup.get_text(" ", strip=True)
        ytm = _extract_pct_after(text, ("MBM", "доходност", "YTM"))
        dur = _extract_after(text, ("дюрац",))
        if ytm is None and dur is None:
            return None
        return MBMValue(
            index_date=date.today(),
            ytm_value=ytm,
            duration=dur,
            source="nbrk_html",
        )


# ────────────────── XLSX parser (вынесен для тестируемости) ──────────────────
def parse_kase_mbm_xlsx(content: bytes) -> List[MBMValue]:
    """Распарсить XLSX-архив MBM с KASE.

    Структура (1-индекс):
        A1: "MBM index 2026-05-01 - 2026-05-07"
        A2..D2: заголовки
        A3..: "06.05.2026" | "1,214.9000" | "2.5100" | "2.1800"

    Возвращает список ``MBMValue``, отсортированный по убыванию даты.
    """
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    out: List[MBMValue] = []
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if row is None or all(c is None for c in row):
            continue
        first = row[0]
        # Заголовочная строка с диапазоном дат — пропускаем.
        if isinstance(first, str) and "MBM" in first and "-" in first and any(ch.isdigit() for ch in first):
            continue
        if isinstance(first, str) and ("Дата" in first or "расчет" in first.lower()):
            header_seen = True
            continue
        d = _coerce_date(first)
        if d is None:
            continue
        idx = _to_float(row[1] if len(row) > 1 else None)
        dur = _to_float(row[2] if len(row) > 2 else None)
        moddur = _to_float(row[3] if len(row) > 3 else None)
        if idx is None and dur is None and moddur is None:
            continue
        out.append(
            MBMValue(
                index_date=d,
                ytm_value=idx,
                duration=dur,
                mod_duration=moddur,
                source="kase_xlsx",
            )
        )
    out.sort(key=lambda v: v.index_date, reverse=True)
    return out


# ────────────────── helpers ──────────────────
def _extract_pct_after(text: str, anchors) -> Optional[float]:
    for anchor in anchors:
        m = re.search(rf"{anchor}[^0-9]{{0,40}}([0-9]+[\.,][0-9]+)\s*%?", text, re.IGNORECASE)
        if m:
            return _to_float(m.group(1))
    return None


def _extract_after(text: str, anchors) -> Optional[float]:
    for anchor in anchors:
        m = re.search(rf"{anchor}[^0-9]{{0,40}}([0-9]+[\.,][0-9]+)", text, re.IGNORECASE)
        if m:
            return _to_float(m.group(1))
    return None


def _to_float(s) -> Optional[float]:
    """Привести значение к float. Поддерживает оба формата: '1,214.9000' и '1 214,90'."""
    if s is None or s == "" or s == "-":
        return None
    if isinstance(s, (int, float)):
        return float(s)
    raw = str(s).strip().replace("\u00A0", " ")
    if not raw:
        return None
    has_comma = "," in raw
    has_dot = "." in raw
    if has_comma and has_dot:
        # "1,214.90" — англ. формат: запятая = разряд, точка = десятичная.
        normalized = raw.replace(",", "").replace(" ", "")
    elif has_comma and not has_dot:
        # "1 214,90" — рус. формат: пробел = разряд, запятая = десятичная.
        normalized = raw.replace(" ", "").replace(",", ".")
    else:
        normalized = raw.replace(" ", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def _coerce_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None
