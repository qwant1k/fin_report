"""Robust parser of incoming TradeReport.xlsx coming from ЧДУ.

Highlights
----------
- Determines column indexes by *header text* (resilient to column-order changes).
- Normalises Kazakhstani number/date formatting via `number_utils`.
- Filters out rows with `Статус != "+"` (ignores cancelled / withdrawn orders).
- Does NOT collapse Разм/К/П rows — каждая строка нужна для построения
  position book (REPO_OPEN / REPO_CLOSE / BUY / SELL).
- Detects ЧДУ from filename / participant code.
- Returns a structured `ParsedTradeFile` with rows + warnings + summary.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

from .classification import (
    DEFAULT_CDU_PREFIXES,
    classify_instrument,
    classify_operation,
    detect_cdu_prefix,
)
from .number_utils import parse_int, parse_kz_date, parse_kz_number, s


# Header aliases — the parser tolerates different XLSX column wordings
HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "deal_number": ("сделка №", "сделка n", "deal"),
    "order_number": ("заявка №", "заявка n", "order"),
    "trade_time": ("время",),
    "kp": ("к/п",),
    "remark": ("примечание",),
    "participant_code": ("код участника",),
    "firm_code": ("код фирмы",),
    "partner_code": ("код партнера",),
    "trade_account": ("торговый счет",),
    "regime_code": ("код режима", "код режима (код)"),
    "instrument_code": ("код инструмента", "код инструмента (код)"),
    "price": ("цена",),
    "lots": ("лоты",),
    "volume": ("объем",),
    "settlement_date": ("дата расчетов",),
    "accrued_interest_volume": ("объем нкд",),
    "yield_pct": ("доходность",),
    "period_code": ("период (код)", "период"),
    "redemption_price": ("цена выкупа",),
    "settlement_code": ("код расчетов",),
    "type_code": ("тип (код)", "тип"),
    "ext_user_code": ("код внешнего пользователя",),
    "commission_total": ("комиссия суммарная",),
    "repo_rate_pct": ("ставка репо, %", "ставка репо"),
    "accrued_interest_volume_repo": ("объем нкд при выкупе",),
    "repo_sum": ("сумма репо",),
    "repo_buyback_sum": ("сумма выкупа репо",),
    "repo_term_days": ("срок репо",),
    "initial_discount_pct": ("начальный дисконт, %",),
    "discount_lower_pct": ("нижний предел дисконта, %",),
    "discount_upper_pct": ("верхний предел дисконта, %",),
    "block_collateral": ("блокировать обеспечение",),
    "commission_clearing": ("комиссия за клиринг",),
    "commission_trading": ("комиссия за торги",),
    "commission_tech": ("комиссия за тех. доступ",),
    "client_code": ("код клиента",),
    "currency_code": ("валюта расчетов (код)", "валюта расчетов"),
    "system_link": ("системная ссылка",),
    "settlement_org": ("расчетная организация (код)", "расчетная организация"),
    "trading_date": ("дата торгов",),
    "clearing_firm_code": ("код клиринговой фирмы",),
    "activity_flag": ("активная/пассивная",),
    "status": ("статус",),
    "nominal_volume": ("объем по номиналу",),
    "clearing_account": ("клиринговый счет",),
    "placement_price": ("цена размещения",),
    "placement_amount": ("сумма",),
    "placement_price_kzt": ("цена размещения, тенге", "цена размещения тенге"),
    "redemption_price_kzt": ("цена выкупа, тенге", "цена выкупа тенге"),
    "securities_to_execute": ("бумаг к исполнению",),
}


@dataclass
class ParsedRow:
    raw_index: int
    fields: Dict[str, Any]


@dataclass
class ParsedTradeFile:
    cdu_prefix: Optional[str]
    cdu_name: Optional[str]
    trade_date: Optional[date]
    rows: List[ParsedRow] = field(default_factory=list)
    skipped: List[Dict[str, Any]] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    sha256: Optional[str] = None
    filename: Optional[str] = None

    @property
    def rows_parsed(self) -> int:
        return len(self.rows)

    @property
    def rows_skipped(self) -> int:
        return len(self.skipped)

    def trade_date_or(self, fallback: date) -> date:
        return self.trade_date or fallback


class TradeReportParser:
    """Parse a single XLSX file (one ЧДУ → list of trades for one date)."""

    def __init__(self, file_path: Path | str, original_name: Optional[str] = None) -> None:
        self.file_path = Path(file_path)
        self.original_name = original_name or self.file_path.name

    # ───────── public API ─────────
    def parse(self) -> ParsedTradeFile:
        wb = load_workbook(self.file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return ParsedTradeFile(cdu_prefix=None, cdu_name=None, trade_date=None,
                                   warnings=["Файл пустой"], filename=self.original_name)

        col_index = self._build_column_index(header_row)

        result = ParsedTradeFile(cdu_prefix=None, cdu_name=None, trade_date=None,
                                 filename=self.original_name)
        result.sha256 = self._sha256()

        # missing-column warning
        required = ["kp", "regime_code", "instrument_code", "status",
                    "participant_code", "trade_account", "volume"]
        for key in required:
            if key not in col_index:
                result.warnings.append(f"Отсутствует колонка: {key} ({HEADER_ALIASES[key][0]})")

        cdu_prefix_votes: Dict[str, int] = {}
        first_trade_date: Optional[date] = None

        for r_idx, raw_row in enumerate(rows_iter, start=2):
            if raw_row is None:
                continue
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw_row):
                continue

            data = self._row_to_dict(raw_row, col_index)

            # фильтр исполненных сделок (+ = Halyk, M = Centras/others)
            status_val = s(data.get("status"))
            if status_val not in ("+", "M"):
                result.skipped.append({"row": r_idx, "reason": f"status={status_val}"})
                continue

            # Operation/category классификация
            regime = s(data.get("regime_code"))
            kp = s(data.get("kp"))
            instrument_code = s(data.get("instrument_code"))
            currency_code = s(data.get("currency_code"))
            clearing_account = s(data.get("clearing_account"))
            data["operation_type"] = classify_operation(
                regime, kp, instrument_code, currency_code, clearing_account
            )
            data["instrument_category"] = classify_instrument(regime, instrument_code)

            # ЧДУ голосование
            participant = s(data.get("participant_code")) or ""
            prefix = detect_cdu_prefix(participant, self.original_name)
            if prefix:
                cdu_prefix_votes[prefix] = cdu_prefix_votes.get(prefix, 0) + 1

            # Trade date — берём дату торгов (или дата расчётов как fallback)
            td = parse_kz_date(data.get("trading_date")) or parse_kz_date(data.get("settlement_date"))
            if td and (first_trade_date is None or td < first_trade_date):
                first_trade_date = td

            data["trade_date"] = td
            data["settlement_date"] = parse_kz_date(data.get("settlement_date"))
            data["trading_date"] = parse_kz_date(data.get("trading_date"))

            # числа
            for k in (
                "price", "lots", "volume", "accrued_interest_volume", "yield_pct",
                "redemption_price", "commission_total", "repo_rate_pct",
                "accrued_interest_volume_repo", "repo_sum", "repo_buyback_sum",
                "initial_discount_pct", "discount_lower_pct", "discount_upper_pct",
                "commission_clearing", "commission_trading", "commission_tech",
                "nominal_volume", "placement_price", "placement_amount",
                "placement_price_kzt", "redemption_price_kzt", "securities_to_execute",
            ):
                data[k] = parse_kz_number(data.get(k))
            data["repo_term_days"] = parse_int(data.get("repo_term_days"))

            # Текстовые
            for k in (
                "deal_number", "order_number", "trade_time", "kp", "participant_code",
                "firm_code", "partner_code", "trade_account", "regime_code",
                "instrument_code", "period_code", "settlement_code", "type_code",
                "client_code", "currency_code", "system_link", "settlement_org",
                "clearing_firm_code", "activity_flag", "status", "clearing_account",
            ):
                if data.get(k) is not None:
                    data[k] = str(data[k]).strip()

            result.rows.append(ParsedRow(raw_index=r_idx, fields=data))

        # Устанавливаем итоговые поля
        if cdu_prefix_votes:
            top = max(cdu_prefix_votes, key=cdu_prefix_votes.get)
            result.cdu_prefix = top
            result.cdu_name = DEFAULT_CDU_PREFIXES.get(top)
        result.trade_date = first_trade_date

        if result.cdu_prefix is None:
            result.warnings.append("Не удалось определить ЧДУ — необходимо выбрать вручную.")
        if result.trade_date is None:
            result.warnings.append("Не удалось определить дату торгов.")

        return result

    # ───────── helpers ─────────
    def _sha256(self) -> str:
        h = hashlib.sha256()
        with self.file_path.open("rb") as f:
            for chunk in iter(lambda: f.read(65536), b""):
                h.update(chunk)
        return h.hexdigest()

    def _build_column_index(self, header_row: Iterable[Any]) -> Dict[str, int]:
        """Map normalised aliases to the actual column index (1-based)."""
        index: Dict[str, int] = {}
        for i, val in enumerate(header_row, start=0):
            if val is None:
                continue
            normalised = re.sub(r"\s+", " ", str(val).strip().lower())
            for key, aliases in HEADER_ALIASES.items():
                if normalised in aliases or any(a in normalised for a in aliases):
                    if key not in index:
                        index[key] = i
                    break
        return index

    def _row_to_dict(self, raw_row: Tuple[Any, ...], col_index: Dict[str, int]) -> Dict[str, Any]:
        out: Dict[str, Any] = {}
        for key, idx in col_index.items():
            if idx < len(raw_row):
                out[key] = raw_row[idx]
        return out
